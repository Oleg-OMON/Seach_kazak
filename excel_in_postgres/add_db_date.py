import openpyxl



wookbook = openpyxl.load_workbook("vgd_baza_pereselencev_20221031.xlsx")
worksheet = wookbook.active
new = []

for i in range(3, 5):
    j = []
    for col in worksheet.iter_cols(1, worksheet.max_column):
        j.append(col[i].value)
    new.append(j)
print(new)


def add_date(list_kazak):
    pass
